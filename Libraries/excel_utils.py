import os
import openpyxl

def get_test_data():
    """Reads the RunManager sheet, then cross-references data from component sheets."""
    base_dir = os.path.dirname(os.path.dirname(__file__))
    excel_path = os.path.join(base_dir, "TestData", "batch_input.xlsx")
    
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Missing Test Data file at: {excel_path}")

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    # 1. Read the Master Sheet (RunManager)
    if "RunManager" not in wb.sheetnames:
        raise ValueError("Excel file must contain a 'RunManager' sheet.")
        
    ws_run = wb["RunManager"]
    run_headers = [cell.value for cell in ws_run[1]]
    
    executable_tcs = []
    
    # Find all TCs marked 'Y'
    for row in ws_run.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(run_headers, row))
        if str(row_dict.get("RunFlag", "")).upper() == 'Y':
            tc_id = row_dict.get("TC_ID")
            
            # 2. Build a unified data object for this TC_ID
            tc_master_data = {
                "TC_ID": tc_id,
                "Description": row_dict.get("Description", "")
            }
            
            # 3. Cross-reference: Hunt through all OTHER sheets for this TC_ID
            for sheet_name in wb.sheetnames:
                if sheet_name == "RunManager":
                    continue # Skip master sheet
                
                ws_component = wb[sheet_name]
                comp_headers = [cell.value for cell in ws_component[1]]
                
                # Add a sub-dictionary for this specific sheet
                tc_master_data[sheet_name] = {}
                
                for comp_row in ws_component.iter_rows(min_row=2, values_only=True):
                    comp_dict = dict(zip(comp_headers, comp_row))
                    
                    # If the TC_ID matches, attach this sheet's data!
                    if comp_dict.get("TC_ID") == tc_id:
                        tc_master_data[sheet_name] = comp_dict
                        break # Found it, move to next sheet
            
            executable_tcs.append(tc_master_data)
            
    return executable_tcs